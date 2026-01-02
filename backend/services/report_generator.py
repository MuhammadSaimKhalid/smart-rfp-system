import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from typing import List, Dict, Optional, Any
from backend.src.agents.rfp_architect import ProposalSchema

# Constants for Styling
HEADER_FILL_COLOR = "4F81BD" # Blue
RFP_HEADER_COLOR = "1F497D" 
VENDOR_FILL_COLORS = ["C0504D", "9BBB59", "8064A2", "4BACC6", "F79646"] # Red, Green, Purple, Blue, Orange
LIGHT_FILL_COLORS = ["F2DCDB", "EBF1DE", "E4DFEC", "DAEEF3", "FDE9D9"]

class ReportGenerator:
    """
    Service to generate dynamic Excel reports for RFP Comparisons.
    Designed to be used as a Tool by AI Agents.
    """
    
    def __init__(self, output_dir: str = "data/reports"):
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)

    def generate_comparison_matrix(self, schema: ProposalSchema, vendors: List[str], vendor_data: Dict[str, Dict[str, Any]] = None) -> str:
        """
        Dynamically generates a Multi-Vendor Comparison Matrix.
        Structure is DERIVED from schema.rfp_headers.
        
        Heuristic: 
        - 'Item', 'Description' -> Fixed Columns (Left)
        - All other headers (Qty, Unit, Price, Total) -> Repeated per Vendor.
        """
        print(f"--- Generating Dynamic Matrix for {len(vendors)} Vendors ---")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bid Analysis"
        
        # --- Styles ---
        header_font = Font(bold=True, size=11, color="FFFFFF")
        cat_font = Font(bold=True, size=11)
        cat_fill = PatternFill("solid", fgColor="DCE6F1")
        
        # --- 1. Determine Column Structure ---
        # Default fallback if LLM is lazy
        discovered_headers = schema.rfp_headers if schema.rfp_headers else ["Item ID", "Description", "Quantity", "Unit", "Unit Price", "Total Cost"]
        
        # Identify Fixed vs Variable
        fixed_headers = []
        variable_headers = []
        
        for h in discovered_headers:
            h_lower = h.lower()
            if "item" in h_lower or "desc" in h_lower or "scope" in h_lower:
                fixed_headers.append(h)
            else:
                variable_headers.append(h)
                
        # If no variable headers found (edge case), default to Qty/Price
        if not variable_headers:
            variable_headers = ["Quantity", "Unit", "Unit Price", "Total Cost"]

        print(f"Structure: Fixed={fixed_headers} | Variable (Per Vendor)={variable_headers}")

        # --- 2. Build Excel Headers ---
        
        # Row 1: Super Headers
        # Fixed Block
        if fixed_headers:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(fixed_headers))
            cell = ws.cell(row=1, column=1, value="RFP SCOPE")
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=RFP_HEADER_COLOR)
            cell.alignment = Alignment(horizontal="center")
            
        # Vendor Blocks
        current_col = len(fixed_headers) + 1
        vendor_starts = {} # map vendor -> start_col
        
        block_width = len(variable_headers)
        
        for idx, vendor in enumerate(vendors):
            vendor_starts[vendor] = current_col
            # Super Header
            ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + block_width - 1)
            cell = ws.cell(row=1, column=current_col, value=vendor.upper())
            cell.font = header_font
            color_idx = idx % len(VENDOR_FILL_COLORS)
            cell.fill = PatternFill("solid", fgColor=VENDOR_FILL_COLORS[color_idx])
            cell.alignment = Alignment(horizontal="center")
            
            current_col += block_width

        # Row 2: Actual Headers
        headers = fixed_headers.copy()
        for vendor in vendors:
            headers.extend(variable_headers)
            
        ws.append(headers)
        
        # Style Row 2
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.font = Font(bold=True)
            cell.border = Border(bottom=Side(style='thick'))
            cell.alignment = Alignment(horizontal="center")
            
            if col_num <= len(fixed_headers):
                cell.fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
                cell.font = Font(bold=True, color="FFFFFF")
            else:
                # Determine vendor index
                # relative col = col_num - len(fixed_headers) - 1
                # v_index = relative // block_width
                rel = col_num - len(fixed_headers) - 1
                v_index = rel // block_width
                color_idx = v_index % len(LIGHT_FILL_COLORS)
                cell.fill = PatternFill("solid", fgColor=LIGHT_FILL_COLORS[color_idx])

        # --- 3. Populate Data ---
        current_row = 3
        
        for category in schema.categories:
            # Category Header
            total_cols = len(fixed_headers) + (len(vendors) * block_width)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
            cell = ws.cell(row=current_row, column=1, value=category.name)
            cell.font = cat_font
            cell.fill = cat_fill
            current_row += 1
            
            cat_start_row = current_row
            
            for item in category.items:
                # Fill Fixed Cols
                if len(fixed_headers) >= 1:
                    ws.cell(row=current_row, column=1, value=item.item_id).alignment = Alignment(horizontal="center")
                if len(fixed_headers) >= 2:
                    ws.cell(row=current_row, column=2, value=item.description).alignment = Alignment(wrap_text=True)
                
                # Fill Vendor Cols
                for vendor in vendors:
                    start_col = vendor_starts[vendor]
                    
                    # Fetch Data
                    v_data = vendor_data.get(vendor, {}).get(item.item_id, {}) if vendor_data else {}
                    
                    # Iterate Variable Headers to map data
                    for i, h_name in enumerate(variable_headers):
                        cell_col = start_col + i
                        key_guess = h_name.lower().replace(" ", "_").replace(".", "") 
                        
                        val = v_data.get(key_guess)
                        
                        # Fallbacks for specific standard fields
                        if val is None:
                            if "qty" in key_guess or "quantity" in key_guess:
                                val = item.quantity or "TBD"
                            elif "unit" == key_guess: 
                                val = item.unit
                            elif "price" in key_guess or "cost" in key_guess:
                                if "total" in key_guess:
                                    # Formula Logic
                                    q_idx = -1
                                    p_idx = -1
                                    for vi, vh in enumerate(variable_headers):
                                        vh_low = vh.lower()
                                        if "qty" in vh_low or "quantity" in vh_low: q_idx = vi
                                        if "price" in vh_low or "rate" in vh_low: p_idx = vi
                                    
                                    if q_idx >= 0 and p_idx >= 0:
                                        q_cell = f"{get_column_letter(start_col + q_idx)}{current_row}"
                                        p_cell = f"{get_column_letter(start_col + p_idx)}{current_row}"
                                        val = f"={q_cell}*{p_cell}"
                                    else:
                                        val = "" 

                        ws.cell(row=current_row, column=cell_col, value=val)

                current_row += 1
                
            # Subtotals (Bottom of Category)
            for vendor in vendors:
                start_col = vendor_starts[vendor]
                for i, h_name in enumerate(variable_headers):
                     if "total" in h_name.lower() or "cost" in h_name.lower() or "amount" in h_name.lower():
                         col_idx = start_col + i
                         col_let = get_column_letter(col_idx)
                         formula = f"=SUM({col_let}{cat_start_row}:{col_let}{current_row-1})"
                         
                         ws.cell(row=current_row, column=col_idx-1, value="Subtotal:").font = Font(bold=True)
                         ws.cell(row=current_row, column=col_idx, value=formula).font = Font(bold=True)

            current_row += 2 

        # --- Borders ---
        self._set_border(ws, f"A1:{get_column_letter(total_cols)}{current_row}")
        
        filename = f"Dynamic_Matrix_{len(vendors)}Vendors.xlsx"
        output_path = os.path.join(self.output_dir, filename)
        wb.save(output_path)
        print(f"Matrix Saved: {output_path}")
        return os.path.abspath(output_path)

    def _set_border(self, ws, cell_range):
        thin = Side(border_style="thin", color="000000")
        for row in ws[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

