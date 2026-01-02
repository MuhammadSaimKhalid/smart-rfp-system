"""
Comparison Matrix Builder

Combines the blank RFP structure with multiple vendor proposals
to create a multi-vendor comparison matrix.
"""

from typing import List, Dict, Any, Optional
from pydantic import BaseModel, Field
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

from backend.src.agents.form_structure_analyzer import (
    ProposalFormStructure,
    DiscoveredFormRow
)
from backend.src.agents.vendor_data_extractor import (
    VendorProposalData,
    FilledFormRow
)


class ComparisonMatrixBuilder:
    """
    Builds multi-vendor comparison matrices from RFP structure and vendor data.
    
    Output: Excel file matching the format of AV - Bid Analysis Spreadsheet.xlsx
    """
    
    def __init__(self):
        self.header_fill = PatternFill(start_color="0066B2", end_color="0066B2", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True)
        self.section_font = Font(bold=True)
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def build_comparison_dataframe(
        self,
        rfp_structure: ProposalFormStructure,
        vendor_proposals: List[VendorProposalData]
    ) -> pd.DataFrame:
        """
        Build a pandas DataFrame for the comparison matrix.
        
        Structure:
        | Item | Description | Vendor1 Qty | Vendor1 Unit | Vendor1 UC | Vendor1 Total | Vendor2 Qty | ...
        """
        rows = []
        
        # Create vendor lookup by item_id for each proposal
        vendor_lookups = {}
        for proposal in vendor_proposals:
            vendor_lookups[proposal.vendor_name] = {
                row.item_id: row for row in proposal.filled_rows
            }
        
        # Track current section for section headers
        current_section = None
        
        for rfp_row in rfp_structure.rows:
            row_data = {}
            
            # Add section header row if section changed
            if rfp_row.section and rfp_row.section != current_section:
                current_section = rfp_row.section
                section_row = {"Item": rfp_row.section, "Description of Work": ""}
                # Add empty vendor columns
                for proposal in vendor_proposals:
                    section_row[f"{proposal.vendor_name} Qty"] = ""
                    section_row[f"{proposal.vendor_name} Unit"] = ""
                    section_row[f"{proposal.vendor_name} Unit Cost"] = ""
                    section_row[f"{proposal.vendor_name} Total"] = ""
                rows.append(section_row)
            
            # Fixed columns (from RFP)
            row_data["Item"] = rfp_row.item_id
            row_data["Description of Work"] = rfp_row.description
            
            # Vendor columns (from each proposal)
            for proposal in vendor_proposals:
                vendor_row = vendor_lookups[proposal.vendor_name].get(rfp_row.item_id)
                
                if vendor_row:
                    row_data[f"{proposal.vendor_name} Qty"] = vendor_row.quantity or rfp_row.values.get("quantity", "")
                    row_data[f"{proposal.vendor_name} Unit"] = vendor_row.unit or rfp_row.values.get("unit", "")
                    row_data[f"{proposal.vendor_name} Unit Cost"] = vendor_row.unit_cost or ""
                    row_data[f"{proposal.vendor_name} Total"] = vendor_row.total or ""
                else:
                    # No vendor data for this item
                    row_data[f"{proposal.vendor_name} Qty"] = rfp_row.values.get("quantity", "")
                    row_data[f"{proposal.vendor_name} Unit"] = rfp_row.values.get("unit", "")
                    row_data[f"{proposal.vendor_name} Unit Cost"] = "—"
                    row_data[f"{proposal.vendor_name} Total"] = "—"
            
            rows.append(row_data)
        
        # Add Grand Total row
        total_row = {"Item": "", "Description of Work": "GRAND TOTAL"}
        for proposal in vendor_proposals:
            total_row[f"{proposal.vendor_name} Qty"] = ""
            total_row[f"{proposal.vendor_name} Unit"] = ""
            total_row[f"{proposal.vendor_name} Unit Cost"] = ""
            total_row[f"{proposal.vendor_name} Total"] = proposal.grand_total or ""
        rows.append(total_row)
        
        return pd.DataFrame(rows)
    
    def build_comparison_excel(
        self,
        rfp_structure: ProposalFormStructure,
        vendor_proposals: List[VendorProposalData],
        output_path: str,
        include_vendor_info: bool = True
    ) -> str:
        """
        Build a formatted Excel comparison matrix.
        
        Args:
            rfp_structure: Discovered RFP form structure
            vendor_proposals: List of extracted vendor data
            output_path: Path for output Excel file
            include_vendor_info: Include vendor contact info header
            
        Returns:
            Path to generated Excel file
        """
        print(f"--- Building Comparison Matrix for {len(vendor_proposals)} vendors ---")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Bid Comparison"
        
        current_row = 1
        
        # Add RFP title
        ws.cell(row=current_row, column=1, value=rfp_structure.form_title)
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
        current_row += 2
        
        if include_vendor_info:
            # Add vendor info header
            col = 3  # Start after Item and Description
            for proposal in vendor_proposals:
                ws.cell(row=current_row, column=col, value=proposal.vendor_name)
                ws.cell(row=current_row, column=col).font = Font(bold=True)
                ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col + 3)
                
                if proposal.vendor_contact:
                    ws.cell(row=current_row + 1, column=col, value=proposal.vendor_contact)
                if proposal.vendor_license:
                    ws.cell(row=current_row + 2, column=col, value=f"License: {proposal.vendor_license}")
                
                col += 4  # Move to next vendor block
            current_row += 4
        
        # Build DataFrame
        df = self.build_comparison_dataframe(rfp_structure, vendor_proposals)
        
        # Add column headers
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=col_name)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = self.thin_border
        current_row += 1
        
        # Add data rows
        for _, row in df.iterrows():
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.border = self.thin_border
                
                # Format section headers
                if col_idx == 1 and str(value).startswith(('I ', 'II ', 'III ', 'IV ', 'V ')):
                    cell.font = self.section_font
                
                # Format currency values
                if isinstance(value, str) and value.startswith('$'):
                    cell.alignment = Alignment(horizontal='right')
                    
            current_row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8    # Item
        ws.column_dimensions['B'].width = 50   # Description
        
        for col in range(3, len(df.columns) + 1):
            ws.column_dimensions[chr(64 + col) if col <= 26 else f"A{chr(64 + col - 26)}"].width = 12
        
        # Save
        wb.save(output_path)
        print(f"  ✓ Saved comparison matrix to: {output_path}")
        
        return output_path
    
    def build_from_selected_proposals(
        self,
        rfp_structure: ProposalFormStructure,
        all_proposals: List[VendorProposalData],
        selected_proposal_ids: List[str],
        output_path: str
    ) -> str:
        """
        Build comparison matrix for only selected/accepted proposals.
        
        Args:
            rfp_structure: RFP form structure
            all_proposals: All vendor proposals (saved in DB)
            selected_proposal_ids: IDs of proposals to include in comparison
            output_path: Output file path
        """
        # Filter to selected proposals only
        selected_proposals = [
            p for p in all_proposals 
            if p.proposal_id in selected_proposal_ids
        ]
        
        if not selected_proposals:
            raise ValueError("No matching proposals found for the selected IDs")
        
        print(f"  Building comparison for {len(selected_proposals)}/{len(all_proposals)} proposals")
        
        return self.build_comparison_excel(
            rfp_structure,
            selected_proposals,
            output_path
        )


# --- Convenience Functions ---

def generate_comparison_report(
    rfp_structure: ProposalFormStructure,
    vendor_proposals: List[VendorProposalData],
    output_dir: str = ".",
    filename: str = "Comparison_Matrix.xlsx"
) -> str:
    """
    Convenience function to generate a comparison report.
    """
    builder = ComparisonMatrixBuilder()
    output_path = os.path.join(output_dir, filename)
    return builder.build_comparison_excel(rfp_structure, vendor_proposals, output_path)


# --- Test ---
if __name__ == "__main__":
    print("=== Testing Comparison Matrix Builder ===\n")
    
    # Create mock data for testing
    from backend.src.agents.form_structure_analyzer import DiscoveredFormRow
    
    # Mock RFP structure
    mock_structure = ProposalFormStructure(
        form_title="AUDUBON VILLAS CONDOMINIUM - REPAIR SPECIFICATIONS",
        tables=[],
        fixed_columns=["Item", "Description of Work"],
        vendor_columns=["Quantity", "Unit", "Unit Cost", "Total"],
        sections=["I Structural", "II Balcony Restoration"],
        rows=[
            DiscoveredFormRow(section="I Structural", item_id="1", description="Wall sheathing repairs and replacement as needed.", values={"quantity": "TBD", "unit": "SF"}),
            DiscoveredFormRow(section="I Structural", item_id="2", description="Wall framing members repairs and replacement as needed.", values={"quantity": "TBD", "unit": "LF"}),
            DiscoveredFormRow(section="II Balcony Restoration", item_id="1", description="Remove and replace all existing ceiling finishes.", values={"quantity": "13,450", "unit": "SF"}),
        ]
    )
    
    # Mock vendor data
    mock_vendors = [
        VendorProposalData(
            proposal_id="prop-001",
            rfp_id="rfp-audubon",
            vendor_name="DueAll",
            filled_rows=[
                FilledFormRow(section="I Structural", item_id="1", description="Wall sheathing repairs", quantity="TBD", unit="SF", unit_cost="$4.10", total="$4.10"),
                FilledFormRow(section="I Structural", item_id="2", description="Wall framing members", quantity="TBD", unit="LF", unit_cost="$7.49", total="$7.49"),
                FilledFormRow(section="II Balcony Restoration", item_id="1", description="Remove and replace ceiling", quantity="13,450", unit="SF", unit_cost="$9.75", total="$131,137.50"),
            ],
            grand_total="$1,122,772.91"
        ),
        VendorProposalData(
            proposal_id="prop-002",
            rfp_id="rfp-audubon",
            vendor_name="IECON",
            filled_rows=[
                FilledFormRow(section="I Structural", item_id="1", description="Wall sheathing repairs", quantity="TBD", unit="SF", unit_cost="$8.00", total="$0"),
                FilledFormRow(section="I Structural", item_id="2", description="Wall framing members", quantity="TBD", unit="LF", unit_cost="$15.00", total="$0"),
                FilledFormRow(section="II Balcony Restoration", item_id="1", description="Remove and replace ceiling", quantity="13,450", unit="SF", unit_cost="$6.50", total="$87,425.00"),
            ],
            grand_total="$989,500.00"
        )
    ]
    
    # Generate comparison
    builder = ComparisonMatrixBuilder()
    output_path = builder.build_comparison_excel(
        mock_structure,
        mock_vendors,
        "Test_Comparison_Matrix.xlsx"
    )
    
    print(f"\nGenerated: {output_path}")
